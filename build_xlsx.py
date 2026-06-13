"""Build a per-car comparison spreadsheet from the two SUMO tripinfo files,
with live Excel AVERAGE formulas at the bottom."""
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SUMO = "sumo/"


def load(path):
    out = {}
    for v in ET.parse(path).getroot().findall("tripinfo"):
        out[v.get("id")] = {
            "depart": float(v.get("depart")),
            "duration": float(v.get("duration")),
            "timeLoss": float(v.get("timeLoss")),
        }
    return out


cur = load(SUMO + "ti_cf.xml")     # current / fixed light
smart = load(SUMO + "ti_ca.xml")   # smart / actuated light
ids = sorted(set(cur) & set(smart), key=lambda i: cur[i]["depart"])

wb = Workbook()
ws = wb.active
ws.title = "Per-car comparison"
ARIAL = "Arial"

ws.merge_cells("A1:F1")
ws["A1"] = "Зүүн дөрвөн зам — every car, current vs smart light"
ws["A1"].font = Font(name=ARIAL, bold=True, size=14)
ws.merge_cells("A2:F2")
ws["A2"] = "Each row is one vehicle. Trip time = when it left the map minus when it entered."
ws["A2"].font = Font(name=ARIAL, italic=True, size=10, color="666666")

headers = ["Vehicle ID", "Enters at (s)", "Current light\ntrip time (s)",
           "Smart light\ntrip time (s)", "Time saved (s)", "Delay cut (s)"]
hrow = 4
blue = PatternFill("solid", fgColor="DDEBF7")
thin = Side(style="thin", color="BBBBBB")
border = Border(bottom=thin, top=thin, left=thin, right=thin)
for c, h in enumerate(headers, 1):
    cell = ws.cell(hrow, c, h)
    cell.font = Font(name=ARIAL, bold=True, size=10)
    cell.fill = blue
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

r = hrow + 1
for i in ids:
    ws.cell(r, 1, i)
    ws.cell(r, 2, cur[i]["depart"])
    ws.cell(r, 3, cur[i]["duration"])
    ws.cell(r, 4, smart[i]["duration"])
    ws.cell(r, 5, f"=C{r}-D{r}")          # time saved
    ws.cell(r, 6, f"=F_cur{r}")           # placeholder, fixed below
    r += 1
last = r - 1

# delay-cut column uses the timeLoss values; write them with a formula vs raw
# (store current/smart timeLoss in helper cols H,I then compute, keeps it transparent)
ws.cell(hrow, 8, "cur delay"); ws.cell(hrow, 9, "smart delay")
ws.cell(hrow, 8).font = Font(name=ARIAL, size=9, color="999999")
ws.cell(hrow, 9).font = Font(name=ARIAL, size=9, color="999999")
for idx, i in enumerate(ids):
    rr = hrow + 1 + idx
    ws.cell(rr, 8, cur[i]["timeLoss"])
    ws.cell(rr, 9, smart[i]["timeLoss"])
    ws.cell(rr, 6, f"=H{rr}-I{rr}")       # delay cut = current delay - smart delay

# averages row
avg = last + 1
ws.cell(avg, 1, "AVERAGE").font = Font(name=ARIAL, bold=True)
ws.cell(avg, 2, f"=AVERAGE(B{hrow+1}:B{last})")
for col in ("C", "D", "E", "F"):
    ws.cell(avg, {"C": 3, "D": 4, "E": 5, "F": 6}[col],
            f"=AVERAGE({col}{hrow+1}:{col}{last})")
yellow = PatternFill("solid", fgColor="FFF2CC")
for c in range(1, 7):
    cell = ws.cell(avg, c)
    cell.font = Font(name=ARIAL, bold=True)
    cell.fill = yellow
    cell.border = Border(top=Side(style="double"))

# number formats + default font on body
for row in ws.iter_rows(min_row=hrow + 1, max_row=avg, min_col=2, max_col=9):
    for cell in row:
        if cell.value is not None:
            cell.number_format = "0.0"
            if not cell.font.bold:
                cell.font = Font(name=ARIAL, size=10)

widths = {"A": 11, "B": 13, "C": 14, "D": 14, "E": 13, "F": 12, "G": 3, "H": 9, "I": 10}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[hrow].height = 30
ws.freeze_panes = "A5"

wb.save("zddz_per_car.xlsx")
print(f"wrote zddz_per_car.xlsx with {len(ids)} cars (rows {hrow+1}-{last}), averages on row {avg}")
